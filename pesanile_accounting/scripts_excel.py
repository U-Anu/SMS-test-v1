import openpyxl
from pesanile_accounting.scripts import create_or_get_asset_type,create_or_get_gl_line,create_or_get_account_type,create_or_get_account_category,create_or_get_account
def get_all_columns_data(file_path):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Iterate over all rows and columns, starting from the first row
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        filtered_row = [str(cell) for cell in row if cell is not None ]
        description = filtered_row[0]
        type_acc_assets = filtered_row[1]
        acc_and_glline = filtered_row[2]
        asset_obj = create_or_get_asset_type(type_acc_assets,description)
        print('asset_obj ',asset_obj)
        glline_obj = create_or_get_gl_line(acc_and_glline,type_acc_assets,description,asset_type=asset_obj)
        print('glline_obj ',glline_obj)
        acc_type_obj =create_or_get_account_type(type_acc_assets,description)
        print('acc_type_obj ',acc_type_obj)
        acc_cat_obj =create_or_get_account_category(type_acc_assets,description)
        print('acc_cat_obj ',acc_cat_obj)
        acc_obj =create_or_get_account(acc_and_glline,gl_line=glline_obj,account_category=acc_cat_obj,account_type=acc_type_obj)
        print('acc_obj ',acc_obj)
        for cell in row:
            print(cell)
        print("-" * 40)  # Separator between rows for clarity

# Example usage:
file_path = '/media/user/DATA FILES/RamWorking/24JunOnwards/GenericAccounting/Chart of accounts.xlsx'  # Replace with your file path

get_all_columns_data(file_path)
